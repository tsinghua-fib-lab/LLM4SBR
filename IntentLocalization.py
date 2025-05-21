import pandas as pd
from sklearn.metrics.pairwise import cosine_similarity
import torch
import os
from tqdm import tqdm
from transformers import BertTokenizer, BertModel
import ast
import torch.nn.functional as F
import torch.nn as nn
import openpyxl


device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(f"Using device: {device}")

tokenizer = BertTokenizer.from_pretrained('../../LLMs/bert-base-uncased')
model = BertModel.from_pretrained("../../LLMs/bert-base-uncased").to(device)

def text_to_embedding(text):
    max_chunk_len = 200
    if len(text) > max_chunk_len:
        text_chunks = [text[i:i + max_chunk_len] for i in range(0, len(text), max_chunk_len)]
        chunks_embeddings = []
        for chunk in text_chunks:
            inputs = tokenizer.encode(chunk, return_tensors="pt").to(device)  # Move inputs to GPU
            with torch.no_grad():
                outputs = model(inputs)
            embeddings = outputs.last_hidden_state.mean(dim=1)
            chunks_embeddings.append(embeddings)
        final_encoding = torch.cat(chunks_embeddings, dim=1)
        return final_encoding
    else:
        tokens = tokenizer(text, return_tensors='pt')
        tokens = {key: val.to(device) for key, val in tokens.items()}  # Move all token tensors to GPU
        with torch.no_grad():
            outputs = model(**tokens)
        embeddings = outputs.last_hidden_state.mean(dim=1)
        return embeddings

def unify_second_dimension(embedding):
    target_dim = 768
    if embedding.size(1) != target_dim:
        linear_layer = torch.nn.Linear(embedding.size(1), target_dim).to(device)  # Move linear layer to GPU
        processed_embedding = linear_layer(embedding)
    else:
        processed_embedding = embedding.clone()
    return processed_embedding

def main():
    type = 'tra'
    xlsx_file_path = "Search_data/ml-1m/{}_session_long.xlsx".format(type)
    workbook = openpyxl.load_workbook(xlsx_file_path)
    sheet = workbook.active
    keys = []
    values = []

    key_column_index = 1
    value_column_index = 2

    for row in sheet.iter_rows(min_row=2, values_only=True):
        keys.append(row[key_column_index - 1])
        values.append(row[value_column_index - 1])

    workbook.close()

    df = pd.DataFrame({'Key': keys, 'Value': values})

    xlsx_file_path2 = "Search_data/ml-1m/movie_name.xlsx"
    workbook_1 = openpyxl.load_workbook(xlsx_file_path2)
    sheet_1 = workbook_1.active
    names = []
    names_column_index = 1
    for row in sheet_1.iter_rows(min_row=2, values_only=True):
        names.append(row[names_column_index - 1])
    workbook_1.close()

    knowledge_df = pd.DataFrame({'name': names})
    df['value_emb'] = df['Value'].apply(lambda x: text_to_embedding(str(x)))
    knowledge_df['name_emb'] = knowledge_df['name'].apply(lambda x: text_to_embedding(str(x)))
    df['value_emb'] = df['value_emb'].apply(lambda x: unify_second_dimension(x))
    knowledge_df['name_emb'] = knowledge_df['name_emb'].apply(lambda x: unify_second_dimension(x))

    print(df.head(5))
    print(knowledge_df.head(5))

    for index, row in tqdm(df.iterrows(), total=len(df), desc="Processing Rows", unit="row"):
        value_embedding = row['value_emb']
        name_embeddings = torch.stack(knowledge_df['name_emb'].tolist()).view(len(knowledge_df), -1).to(device)
        similarities = cosine_similarity(value_embedding.cpu().detach().numpy(), name_embeddings.cpu().detach().numpy())
        top5_indices = torch.topk(torch.tensor(similarities.flatten()), 5)[1]
        top5_name = knowledge_df['name'].iloc[top5_indices].tolist()
        df.at[index, 'top5_names'] = str(top5_name)
        top5_name_embs = knowledge_df['name_emb'].iloc[top5_indices].tolist()
        similarities_tensor = torch.tensor(similarities.flatten(), device=device)
        updated_value_emb = sum([emb * sim for emb, sim in zip(top5_name_embs, similarities_tensor)]) / 5
        df.at[index, 'long_emb'] = str(updated_value_emb.cpu().tolist())  # Move to CPU for storage

    print(df.head(5))
    df.to_excel('Search_data/ml-1m/{}_session_long_emb.xlsx'.format(type))

if __name__ == '__main__':
    main()